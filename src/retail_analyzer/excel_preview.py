"""
Embed image thumbnails next to image URL columns and turn URL cells into hyperlinks.

Excel cannot show a picture inside a hover tooltip via openpyxl; we add a dedicated
"Image preview" column and use hyperlink ScreenTips on the URL cell (hover shows text).
"""
from __future__ import annotations

import io
import logging
from pathlib import Path

import pandas as pd
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from PIL import Image

from retail_analyzer.config import AnalyzerConfig
from retail_analyzer.excel_io import parse_image_urls, resolve_image_path
from retail_analyzer.image_fetch import fetch_image_pil

logger = logging.getLogger(__name__)

_HYPERLINK_FONT = Font(color="0563C1", underline="single")
_TOOLTIP = (
    "Click to open this image. A thumbnail is shown in the next column (Image preview)."
)


def _header_column_index(ws, name: str) -> int | None:
    for c in range(1, (ws.max_column or 0) + 1):
        v = ws.cell(1, c).value
        if v is not None and str(v).strip() == str(name).strip():
            return c
    return None


def _resize_for_thumb(im: Image.Image, max_px: int) -> Image.Image:
    w, h = im.size
    if w <= max_px and h <= max_px:
        return im
    scale = min(max_px / w, max_px / h)
    nw, nh = max(1, int(w * scale)), max(1, int(h * scale))
    try:
        resample = Image.Resampling.LANCZOS
    except AttributeError:
        resample = Image.LANCZOS  # type: ignore[attr-defined]
    return im.resize((nw, nh), resample)


def _hyperlink_target(first_ref: str) -> str | None:
    s = first_ref.strip()
    if not s:
        return None
    if s.lower().startswith("http"):
        return s
    p = resolve_image_path(s)
    if p is not None and p.is_file():
        return Path(p).resolve().as_uri()
    return None


def enrich_workbook_with_image_previews(
    wb,
    ws,
    df: pd.DataFrame,
    image_col: str,
    *,
    config: AnalyzerConfig | None = None,
    max_thumb_px: int = 120,
) -> None:
    """
    Insert an "Image preview" column immediately after ``image_col``, embed thumbnails,
    and add hyperlinks + ScreenTips on URL cells. No-op if the column is missing.
    """
    if len(df) == 0:
        return
    col_idx = _header_column_index(ws, image_col)
    if col_idx is None:
        return

    cfg = config or AnalyzerConfig()
    ws.insert_cols(col_idx + 1)
    preview_col = col_idx + 1
    ws.cell(1, preview_col).value = "Image preview"
    prev_letter = get_column_letter(preview_col)
    ws.column_dimensions[prev_letter].width = 18

    for i in range(len(df)):
        row = i + 2
        raw = df.iloc[i][image_col]
        urls = parse_image_urls(raw)
        if not urls:
            continue
        first = urls[0]
        target = _hyperlink_target(first)
        url_cell = ws.cell(row=row, column=col_idx)
        if target:
            ref = f"{get_column_letter(col_idx)}{row}"
            url_cell.hyperlink = Hyperlink(ref=ref, target=target, tooltip=_TOOLTIP)
            url_cell.font = _HYPERLINK_FONT

        img = fetch_image_pil(first, cfg)
        if img is None:
            continue
        img = _resize_for_thumb(img, max_thumb_px)
        bio = io.BytesIO()
        img.save(bio, format="PNG")
        bio.seek(0)
        xl_img = XLImage(bio)
        xl_img.anchor = f"{prev_letter}{row}"
        ws.add_image(xl_img)
        rd = ws.row_dimensions[row]
        cur = rd.height
        want = 96.0
        if cur is None or cur < want:
            rd.height = want
