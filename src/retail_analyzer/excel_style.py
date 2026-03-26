from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def apply_highlights(path: Path, df: pd.DataFrame) -> None:
    """
    Add fills to rows with visual duplicate groups or anomaly flags.
    Expects path to match df (same columns, written without index).
    """
    path = Path(path)
    wb = load_workbook(path)
    ws = wb.active
    if ws is None:
        return

    fill_dup = PatternFill(fill_type="solid", start_color="CCE5FF", end_color="CCE5FF")
    fill_anom = PatternFill(fill_type="solid", start_color="FFE5CC", end_color="FFE5CC")
    fill_both = PatternFill(fill_type="solid", start_color="E5CCFF", end_color="E5CCFF")

    max_col = ws.max_column or 1
    gid_col = "duplicate_group_id" if "duplicate_group_id" in df.columns else "similar_image_group_id"
    has_gid = gid_col in df.columns
    has_flags = "anomaly_flags" in df.columns
    has_flag_col = "anomaly_flag" in df.columns

    for data_i in range(len(df)):
        row_idx = data_i + 2
        gid_v = -1
        if has_gid:
            v = df.iloc[data_i][gid_col]
            try:
                gid_v = int(v)
            except (TypeError, ValueError):
                gid_v = -1
        flag_s = ""
        if has_flags:
            raw = df.iloc[data_i]["anomaly_flags"]
            if raw is not None and not (isinstance(raw, float) and pd.isna(raw)):
                flag_s = str(raw).strip()
        anom_yes = False
        if has_flag_col:
            raw = df.iloc[data_i]["anomaly_flag"]
            if raw is not None and not (isinstance(raw, float) and pd.isna(raw)):
                anom_yes = bool(raw) if not isinstance(raw, str) else str(raw).strip().lower() in (
                    "yes",
                    "true",
                    "1",
                )

        has_dup = gid_v >= 0
        has_anom = anom_yes or bool(flag_s)
        if not has_dup and not has_anom:
            continue
        fill = fill_both if has_dup and has_anom else (fill_dup if has_dup else fill_anom)
        for c in range(1, max_col + 1):
            ws.cell(row=row_idx, column=c).fill = fill

    wb.save(path)
