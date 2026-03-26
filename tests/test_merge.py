import io
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from retail_analyzer.merge import dataframe_to_excel_bytes, merge_visual_duplicate_rows


def test_merge_combines_group():
    df = pd.DataFrame(
        {
            "sku": ["a", "a", "b"],
            "img": ["http://x/1", "http://x/2", "http://y/1"],
            "similar_image_group_id": [0, 0, -1],
        }
    )
    out = merge_visual_duplicate_rows(df, "img", drop_group_id_column=True)
    assert len(out) == 2
    merged_row = out[out["img"].astype(str).str.contains("http://x/2", na=False)].iloc[0]
    assert "http://x/1" in merged_row["img"] and "http://x/2" in merged_row["img"]


def test_merge_keeps_singleton():
    df = pd.DataFrame(
        {
            "sku": ["b"],
            "img": ["http://y/1"],
            "similar_image_group_id": [-1],
        }
    )
    out = merge_visual_duplicate_rows(df, "img", drop_group_id_column=True)
    assert len(out) == 1


def test_excel_export_adds_image_preview_column():
    root = Path(__file__).resolve().parents[1]
    png = root / "tests" / "fixtures" / "rotated_duplicate_dataset" / "product_a_rot_0.png"
    if not png.is_file():
        pytest.skip("rotated_duplicate_dataset fixture not present")
    rel = str(png.relative_to(root)).replace("\\", "/")
    df = pd.DataFrame({"image_urls": [rel], "note": [1]})
    raw = dataframe_to_excel_bytes(df, image_column="image_urls")
    wb = load_workbook(io.BytesIO(raw))
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, (ws.max_column or 0) + 1)]
    assert "Image preview" in headers
